import openpyxl
import pyttsx3
import speech_recognition as sr
import time
import pandas as pd


class Questions:

    def display_message(questionList):
        filename = questionList+".xlsx"
        questions = pd.read_excel(filename)
        
    
    def ask_questions(question_count, user_score, running,questions,recognizer,daily_quota,engine,answers):
        # global question_count, user_score, running

        while running and question_count < daily_quota:
            time.sleep(20)  # Wait for 60 seconds before asking the next question
            timeout = 10  # Set a timeout of 10 seconds

            while True:
                question_text = questions[question_count]
                engine.say(f"Question {question_count + 1}: {question_text}")
                engine.runAndWait()
                engine.say("You can answer by speaking ")

                # Listen for the user's answer
                with sr.Microphone() as source:
                    print(f"Question {question_count + 1}: {question_text}")
                    engine.say("Now, please tell me your answer.")
                    engine.runAndWait()

                    audio = recognizer.listen(source)

                try:
                    user_response = recognizer.recognize_google(audio).lower()
                except sr.UnknownValueError:
                    user_response = ""  # If the speech recognition couldn't understand the response

                # # Allow text input as well
                # if not user_response:
                #     user_response = input("You can also type your answer and press Enter: ")

                # Check the answer
                correct_answer = answers[question_count]
                if user_response == correct_answer.lower():
                    user_score += 1
                    engine.say("That's great, it's the right answer!")
                    engine.runAndWait()
                    question_count += 1
                    break  # Move to the next question
                else:
                    engine.say(f"I'm sorry, that's not quite right. The correct answer is: {correct_answer}")
                    engine.say("Let's try the same question again.")
                    engine.runAndWait()

            # Provide the user's current score
            engine.say(f"Your current score is {user_score} out of {question_count}.")

    def start_questions(questionList):
        filename = questionList+".xlsx"
        # Load the Excel file and extract questions and answers
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        # Extract questions and answers from Excel
        questions = [row[0].value for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=1)]  # Assuming 10 questions
        answers = [row[0].value for row in sheet.iter_rows(min_row=1, max_row=10, min_col=2, max_col=2)]  # Assuming 10 answers

        # Initialize text-to-speech engine
        engine = pyttsx3.init()

        # Initialize speech recognition
        recognizer = sr.Recognizer()

        # Set up the loop
        question_count = 0
        daily_quota = 2

        # Initialize the user's score
        user_score = 0

        # Flag for controlling the loop
        running = False
        
        if not running:
            engine.say("Questionnaire started.")
            engine.runAndWait()
            running = True
            Questions.ask_questions(question_count, user_score, running,questions,recognizer,daily_quota,engine,answers)
        else:
            engine.say("Questionnaire is already running.")
            engine.runAndWait()

    def stop_questions():
        global running
        engine = pyttsx3.init()
        if running:
            engine.say("Questionnaire stopped.")
            engine.runAndWait()
            running = False
            return True
        else:
            engine.say("Questionnaire is not running.")
            engine.runAndWait()
            return False


